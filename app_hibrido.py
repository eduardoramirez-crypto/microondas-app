# --- VERSIÓN HÍBRIDA: APLICACIÓN ORIGINAL ADAPTADA PARA LINUX ---
# Mantiene toda la funcionalidad original pero usa pandas/openpyxl en lugar de xlwings/win32com
# --- FIN DEL COMENTARIO ---

import os
import time
import pandas as pd
import numpy as np
from flask import Flask, request, send_file, render_template_string, redirect, url_for, after_this_request, jsonify, send_from_directory
from werkzeug.utils import secure_filename
import sys
import re
import textwrap
import unicodedata
import glob
import subprocess
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import tempfile
import shutil

app = Flask(__name__)

# Configuración para archivos
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'png', 'jpg', 'jpeg', 'gif', 'bmp'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Crear directorio de uploads si no existe
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def normaliza_na(valor):
    if isinstance(valor, str) and valor.strip().lower() == "n/a":
        return "N/A"
    elif pd.isna(valor):
        return "N/A"
    elif valor == "" or (isinstance(valor, str) and valor.strip() == ""):
        return "N/A"
    return valor

def normaliza_texto(texto):
    if pd.isna(texto):
        return ""
    texto_str = str(texto).strip()
    if texto_str == "" or texto_str.lower() == "nan":
        return ""
    return texto_str

# Obtener el directorio base de la aplicación
if getattr(sys, 'frozen', False):
    base_dir = os.path.dirname(sys.executable)
else:
    base_dir = os.path.dirname(os.path.abspath(__file__))

print(f"Directorio base: {base_dir}")
print("Archivos en el directorio:", os.listdir(base_dir))

# Buscar archivos Excel en el directorio
def encontrar_archivos_excel():
    archivos_excel = []
    for archivo in os.listdir(base_dir):
        if archivo.endswith(('.xlsx', '.xls')):
            archivos_excel.append(archivo)
    return archivos_excel

# Función para crear archivo Excel con pandas/openpyxl
def crear_excel_con_pandas(df, filename, sheet_name="Datos"):
    try:
        # Crear un nuevo workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Escribir los datos del DataFrame
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Aplicar estilos básicos
        for cell in ws[1]:  # Header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Guardar el archivo
        wb.save(filename)
        return True
    except Exception as e:
        print(f"Error al crear Excel: {e}")
        return False

@app.route('/')
def index():
    return send_from_directory('.', 'login.html')

@app.route('/ptpFangio')
def ptp_fangio():
    return send_from_directory('.', 'ptpFangio.html')

@app.route('/ptmpFangio')
def ptmp_fangio():
    return send_from_directory('.', 'ptmpFangio.html')

@app.route('/llenado-automatico')
def llenado_automatico():
    archivos_excel = encontrar_archivos_excel()
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Llenado Automático - Fangio Telecom</title>
        <style>
            body {{
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                margin: 0;
                padding: 20px;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                min-height: 100vh;
                color: #333;
            }}
            .container {{
                max-width: 800px;
                margin: 0 auto;
                background: white;
                border-radius: 15px;
                box-shadow: 0 10px 30px rgba(0,0,0,0.2);
                overflow: hidden;
            }}
            .header {{
                background: linear-gradient(135deg, #2c3e50 0%, #34495e 100%);
                color: white;
                padding: 30px;
                text-align: center;
            }}
            .header h1 {{
                margin: 0;
                font-size: 2.5em;
                font-weight: 300;
            }}
            .content {{
                padding: 40px;
            }}
            .file-section {{
                margin-bottom: 30px;
            }}
            .file-section h2 {{
                color: #2c3e50;
                border-bottom: 3px solid #3498db;
                padding-bottom: 10px;
                margin-bottom: 20px;
            }}
            .file-list {{
                list-style: none;
                padding: 0;
            }}
            .file-item {{
                background: #f8f9fa;
                border: 1px solid #e9ecef;
                border-radius: 8px;
                padding: 15px;
                margin-bottom: 10px;
                display: flex;
                justify-content: space-between;
                align-items: center;
                transition: all 0.3s ease;
            }}
            .file-item:hover {{
                background: #e9ecef;
                transform: translateY(-2px);
                box-shadow: 0 4px 8px rgba(0,0,0,0.1);
            }}
            .file-name {{
                font-weight: 600;
                color: #495057;
            }}
            .btn {{
                background: linear-gradient(135deg, #3498db 0%, #2980b9 100%);
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                cursor: pointer;
                text-decoration: none;
                display: inline-block;
                transition: all 0.3s ease;
                font-weight: 600;
            }}
            .btn:hover {{
                background: linear-gradient(135deg, #2980b9 0%, #1f5f8b 100%);
                transform: translateY(-1px);
                box-shadow: 0 4px 8px rgba(0,0,0,0.2);
            }}
            .back-btn {{
                background: linear-gradient(135deg, #95a5a6 0%, #7f8c8d 100%);
                margin-top: 20px;
            }}
            .back-btn:hover {{
                background: linear-gradient(135deg, #7f8c8d 0%, #6c7b7d 100%);
            }}
            .no-files {{
                text-align: center;
                color: #6c757d;
                font-style: italic;
                padding: 40px;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>📊 Llenado Automático</h1>
                <p>Sistema de Gestión Fangio Telecom - Versión Linux</p>
            </div>
            <div class="content">
                <div class="file-section">
                    <h2>📁 Archivos Excel Disponibles</h2>
                    {f'<ul class="file-list">{"".join([f"<li class=\"file-item\"><span class=\"file-name\">{archivo}</span><a href=\"/procesar/{archivo}\" class=\"btn\">Procesar</a></li>" for archivo in archivos_excel])}</ul>' if archivos_excel else '<div class="no-files">No se encontraron archivos Excel en el directorio</div>'}
                </div>
                <a href="/" class="btn back-btn">← Volver al Inicio</a>
            </div>
        </div>
    </body>
    </html>
    """
    
    return render_template_string(html_content)

@app.route('/procesar/<filename>')
def procesar_archivo(filename):
    try:
        file_path = os.path.join(base_dir, filename)
        
        # Leer el archivo Excel
        df = pd.read_excel(file_path)
        
        # Mostrar información del archivo
        info_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Procesando {filename} - Fangio Telecom</title>
            <style>
                body {{
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    margin: 0;
                    padding: 20px;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    min-height: 100vh;
                    color: #333;
                }}
                .container {{
                    max-width: 1000px;
                    margin: 0 auto;
                    background: white;
                    border-radius: 15px;
                    box-shadow: 0 10px 30px rgba(0,0,0,0.2);
                    overflow: hidden;
                }}
                .header {{
                    background: linear-gradient(135deg, #27ae60 0%, #2ecc71 100%);
                    color: white;
                    padding: 30px;
                    text-align: center;
                }}
                .content {{
                    padding: 40px;
                }}
                .info-grid {{
                    display: grid;
                    grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
                    gap: 20px;
                    margin-bottom: 30px;
                }}
                .info-card {{
                    background: #f8f9fa;
                    border-radius: 10px;
                    padding: 20px;
                    border-left: 4px solid #3498db;
                }}
                .info-card h3 {{
                    margin-top: 0;
                    color: #2c3e50;
                }}
                .btn {{
                    background: linear-gradient(135deg, #3498db 0%, #2980b9 100%);
                    color: white;
                    border: none;
                    padding: 12px 24px;
                    border-radius: 5px;
                    cursor: pointer;
                    text-decoration: none;
                    display: inline-block;
                    margin: 5px;
                    transition: all 0.3s ease;
                    font-weight: 600;
                }}
                .btn:hover {{
                    background: linear-gradient(135deg, #2980b9 0%, #1f5f8b 100%);
                    transform: translateY(-1px);
                }}
                .success {{
                    background: linear-gradient(135deg, #27ae60 0%, #2ecc71 100%);
                }}
                .success:hover {{
                    background: linear-gradient(135deg, #2ecc71 0%, #27ae60 100%);
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>✅ Archivo Procesado: {filename}</h1>
                    <p>Información del archivo Excel - Versión Linux</p>
                </div>
                <div class="content">
                    <div class="info-grid">
                        <div class="info-card">
                            <h3>📊 Dimensiones</h3>
                            <p><strong>Filas:</strong> {len(df)}</p>
                            <p><strong>Columnas:</strong> {len(df.columns)}</p>
                        </div>
                        <div class="info-card">
                            <h3>📋 Columnas</h3>
                            <p><strong>Total:</strong> {len(df.columns)} columnas</p>
                            <p><strong>Primeras 5:</strong> {', '.join(df.columns[:5].tolist())}</p>
                        </div>
                        <div class="info-card">
                            <h3>📁 Información del Archivo</h3>
                            <p><strong>Nombre:</strong> {filename}</p>
                            <p><strong>Tamaño:</strong> {os.path.getsize(file_path) / 1024:.1f} KB</p>
                        </div>
                    </div>
                    <div style="text-align: center;">
                        <a href="/llenado-automatico" class="btn">← Volver a Archivos</a>
                        <a href="/" class="btn success">🏠 Ir al Inicio</a>
                    </div>
                </div>
            </div>
        </body>
        </html>
        """
        
        return render_template_string(info_html)
        
    except Exception as e:
        error_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Error - Fangio Telecom</title>
            <style>
                body {{
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    margin: 0;
                    padding: 20px;
                    background: linear-gradient(135deg, #e74c3c 0%, #c0392b 100%);
                    min-height: 100vh;
                    color: white;
                }}
                .container {{
                    max-width: 600px;
                    margin: 0 auto;
                    background: rgba(255,255,255,0.1);
                    border-radius: 15px;
                    padding: 40px;
                    text-align: center;
                    backdrop-filter: blur(10px);
                }}
                .btn {{
                    background: rgba(255,255,255,0.2);
                    color: white;
                    border: 1px solid rgba(255,255,255,0.3);
                    padding: 12px 24px;
                    border-radius: 5px;
                    text-decoration: none;
                    display: inline-block;
                    margin: 10px;
                    transition: all 0.3s ease;
                }}
                .btn:hover {{
                    background: rgba(255,255,255,0.3);
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <h1>❌ Error al Procesar</h1>
                <p><strong>Archivo:</strong> {filename}</p>
                <p><strong>Error:</strong> {str(e)}</p>
                <a href="/llenado-automatico" class="btn">← Volver</a>
                <a href="/" class="btn">🏠 Inicio</a>
            </div>
        </body>
        </html>
        """
        return render_template_string(error_html)

# Rutas adicionales de tu aplicación original
@app.route('/site_survey_checkboxes', methods=['GET'])
def site_survey_checkboxes():
    return send_from_directory('.', 'site_survey_checkboxes.html')

@app.route('/diseno_solucion', methods=['GET', 'POST'])
def diseno_solucion():
    return send_from_directory('.', 'formulario_archivos.html')

@app.route('/site_survey', methods=['GET'])
def site_survey():
    return send_from_directory('.', 'site_survey_checkboxes.html')

@app.route('/seleccion_tipo_llenado')
def seleccion_tipo_llenado():
    return send_from_directory('.', 'seleccion_tipo_llenado.html')

@app.route('/seleccion_llenado_ptp')
def seleccion_llenado_ptp():
    return send_from_directory('.', 'seleccion_llenado_ptp.html')

@app.route('/seleccion_llenado_ptmp')
def seleccion_llenado_ptmp():
    return send_from_directory('.', 'seleccion_llenado_ptmp.html')

@app.route('/subir_imagenes_ptp', methods=['GET', 'POST'])
def subir_imagenes_ptp():
    return send_from_directory('.', 'subir_imagenes_ptp.html')

@app.route('/subir_imagenes_ptp_planos_b', methods=['GET', 'POST'])
def subir_imagenes_ptp_planos_b():
    return send_from_directory('.', 'subir_imagenes_ptp_planos_b.html')

@app.route('/subir_imagenes_ptp_fotos_a', methods=['GET', 'POST'])
def subir_imagenes_ptp_fotos_a():
    return send_from_directory('.', 'subir_imagenes_ptp_fotos_a.html')

@app.route('/subir_imagenes_ptp_fotos_b', methods=['GET', 'POST'])
def subir_imagenes_ptp_fotos_b():
    return send_from_directory('.', 'subir_imagenes_ptp_fotos_b.html')

@app.route('/<path:filename>')
def serve_file(filename):
    return send_from_directory('.', filename)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True) 